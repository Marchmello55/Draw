import asyncio
from openpyxl import Workbook, load_workbook

FILE_NAME = "example.xlsx"

# ---------- СИНХРОННЫЕ ВНУТРЕННИЕ ФУНКЦИИ ----------

def _create_excel_sync():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Users"
    ws1.append(["tg-id", "username", "Телефон"])
    ws2 = wb.create_sheet("Codes")
    ws2.append(["tg-id", "code"])

    wb.save(FILE_NAME)
    print("Файл создан:", FILE_NAME)


def _add_row_sync(sheet_name, row_data):
    wb = load_workbook(FILE_NAME)
    ws = wb[sheet_name]
    ws.append(row_data)
    wb.save(FILE_NAME)


def _update_row_sync(sheet_name, search_column, search_value, update_data: dict):
    """
    Обновляем строку только если в целевом столбце пусто.
    """
    wb = load_workbook(FILE_NAME)
    ws = wb[sheet_name]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    updated = False

    for row in ws.iter_rows(min_row=2):
        if row[headers[search_column]-1].value == search_value:
            for col_name, new_val in update_data.items():
                col_idx = headers[col_name]
                current_val = ws.cell(row=row[0].row, column=col_idx).value
                # обновляем только если пусто
                if current_val in (None, ""):
                    ws.cell(row=row[0].row, column=col_idx, value=new_val)
                    updated = True
            break

    wb.save(FILE_NAME)
    return updated


def _find_row_sync(sheet_name, column_name, value):
    wb = load_workbook(FILE_NAME)
    ws = wb[sheet_name]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[headers[column_name]-1] == value:
            return row
    return None

def _find_all_rows_sync(sheet_name, column_name, value):
    """Найти все строки с нужным значением"""
    wb = load_workbook(FILE_NAME)
    ws = wb[sheet_name]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    col_idx = headers[column_name]

    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx-1] == value:
            result.append(row)
    return result


# ---------- АСИНХРОННЫЕ ОБЁРТКИ ----------

async def create_excel():
    await asyncio.to_thread(_create_excel_sync)

async def add_row(sheet_name: str, row_data: list):
        await asyncio.to_thread(_add_row_sync, sheet_name, row_data)

async def update_row(sheet_name, search_column, search_value, update_data):
    await asyncio.to_thread(_update_row_sync, sheet_name, search_column, search_value, update_data)

async def find_row(sheet_name, column_name, value):
    return await asyncio.to_thread(_find_row_sync, sheet_name, column_name, value)

async def find_all_rows(sheet_name, column_name, value):
    return await asyncio.to_thread(_find_all_rows_sync, sheet_name, column_name, value)


# ---------- ЛОГИКА С ПРОВЕРКАМИ ----------

async def main():
    # Проверяем наличие файла

    await create_excel()

    # Проверяем строку в Users
    await add_row("Users", [1, "qwerty", 234567], "tg-id", 1)


    # Проверяем строку в Codes
    code = await find_row("Codes", "tg-id", 1)
    if not code:
        await add_row("Codes", [1, "qwert"], "code", "qwert")
        print("Добавлен новый код")
    else:
        await update_row("Codes", "tg-id", 1, {"code": "new_code"})
        print("Обновлён код:", code)

    # Итоговый поиск
    print("Users[1]:", await find_row("Users", "tg-id", 1))
    print("Codes[1]:", await find_row("Codes", "tg-id", 1))


if __name__ == "__main__":
    asyncio.run(main())
